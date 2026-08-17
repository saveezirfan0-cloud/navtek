"""
Navtek eOrder parser — reference implementation
================================================

Reads a Teletrac Navman eOrder (.xlsx) and returns a flat dict ready to write
to the "TN Orders 2026/2025/2024" monday.com board (id 5834171978).

Validated against five real eOrders covering five order types — New Business,
Hardware Only (Replacement), Change of Ownership, Service Only Renewal and a
multi-site Add-On. All five parse with identical code; no branching by order
type. See expected/verify.py, which is the acceptance gate for extraction.

    pip install openpyxl
    python eorder_parser.py <file.xlsx>

Two things in here are load-bearing. Read them before refactoring:

1. load_workbook_tolerant()
   TN's Salesforce template emits a <customWorkbookViews> block whose GUID is
   lowercase. openpyxl validates that attribute against an uppercase-only regex
   and hard-fails the whole file. Excel itself opens it fine, so this is a
   parser-strictness problem, not a corrupt file. We strip the block in memory.
   Do not "fix" this by asking the team to re-save the file.

2. find()
   Every value is located by searching for its LABEL, then reading the cell at a
   given offset. Do NOT switch this to hardcoded cell refs like ws['E19'].
   TN revises this template periodically; label-anchored lookups survive a row
   being inserted, hardcoded refs break silently and nobody notices for weeks.

Author: prepared for Saveez, Aug 2026
"""

import io
import json
import re
import sys
import zipfile

import openpyxl

# --------------------------------------------------------------------------
# Loading
# --------------------------------------------------------------------------

def load_workbook_tolerant(source):
    """Open a TN eOrder, repairing the non-standard workbook GUID in memory.

    `source` may be a path or a file-like object (e.g. the response body from
    monday's asset public_url).
    """
    buf = io.BytesIO()
    zin = zipfile.ZipFile(source)
    zout = zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED)
    for item in zin.infolist():
        data = zin.read(item.filename)
        if item.filename == "xl/workbook.xml":
            data = re.sub(
                rb"<customWorkbookViews>.*?</customWorkbookViews>", b"", data, flags=re.S
            )
        zout.writestr(item, data)
    zout.close()
    return openpyxl.load_workbook(buf, data_only=True)


# --------------------------------------------------------------------------
# Label-anchored cell lookup
# --------------------------------------------------------------------------

EMPTY = (None, "", "--Select--", "Select", "---Select---", "#REF!")


def norm(s):
    """Normalise a label for comparison.

    Order matters: strip whitespace FIRST, then the required-field asterisk,
    then strip again. The eOrder contains labels like 'Postcode*  ' with
    trailing spaces after the asterisk — rstrip('*') alone silently misses them.
    """
    return str(s).strip().rstrip("*").strip().lower()


def find(ws, label, dx=1, dy=0, within=None, col_range=None):
    """Find `label` in the sheet, return the value dx columns / dy rows away.

    within    : optional (min_row, max_row) to disambiguate repeated labels.
    col_range : optional (min_col, max_col), 1-indexed. REQUIRED for the contact
                blocks — the eOrder puts General Contact and Accounts Contact
                side by side on the SAME rows (A34 'Name*' vs C34 'Name*'), so a
                row window alone cannot tell them apart.

    Widen these windows if TN adds rows or columns to the template.
    """
    target = norm(label)
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if norm(cell.value) != target:
                continue
            if within and not (within[0] <= cell.row <= within[1]):
                continue
            if col_range and not (col_range[0] <= cell.column <= col_range[1]):
                continue
            value = ws.cell(row=cell.row + dy, column=cell.column + dx).value
            if value not in EMPTY:
                return str(value).strip()
    return None


def find_right(ws, label, max_span=4, **kw):
    """As find(), but scans several columns right until it hits a real value.

    The eOrder is inconsistent about whether a value sits 1, 2 or 3 columns
    from its label depending on the sheet. This absorbs that.
    """
    for dx in range(1, max_span + 1):
        v = find(ws, label, dx=dx, **kw)
        if v is not None:
            return v
    return None


# --------------------------------------------------------------------------
# Extraction
# --------------------------------------------------------------------------

def parse(source):
    wb = load_workbook_tolerant(source)

    gi = wb["General Information"]
    di = wb["Delivery & Install Details"]
    bi = wb["Billing, Order & SIM Activation"]

    data = {
        # --- match key: joins to monday column `order__` (OPPORTUNITY ID #) ---
        "opportunity_id": find_right(gi, "Opportunity Field ID"),

        # --- customer ---
        "company":  find_right(gi, "Company Name", within=(16, 19)),
        "abn":      find_right(gi, "ABN Number"),
        "address":  find_right(gi, "Address", within=(23, 25)),
        "suburb":   find_right(gi, "Suburb", within=(25, 27)),
        "state":    find_right(gi, "State", within=(25, 27)),
        "postcode": find_right(gi, "Postcode", within=(27, 29)),

        # --- site contact: what the installer actually needs ---
        "site_contact_name":  find_right(di, "Contact", within=(14, 16)),
        "site_contact_phone": find_right(di, "Phone No", within=(15, 17)),
        "site_contact_email": find_right(di, "Email", within=(16, 18)),

        # --- general contact (cols A/B) vs accounts contact (cols C/D):
        #     same rows, different columns — col_range is what separates them ---
        "general_name":   find_right(gi, "Name",          within=(33, 35), col_range=(1, 2)),
        "general_email":  find_right(gi, "Email Address", within=(35, 37), col_range=(1, 2)),
        "general_phone":  find_right(gi, "Phone Number",  within=(37, 39), col_range=(1, 2)),
        "accounts_name":  find_right(gi, "Name",          within=(33, 35), col_range=(3, 4)),
        "accounts_email": find_right(gi, "Email Address", within=(35, 37), col_range=(3, 4)),
        "accounts_phone": find_right(gi, "Phone Number",  within=(37, 39), col_range=(3, 4)),

        # --- install ---
        "tn_organising_install": find_right(di, "Is Teletrac Navman Organising Install?"),
        "installer_company":     find_right(di, "Company Name", within=(16, 18)),
        "installer_contact":     find_right(di, "Contact Name", within=(18, 20)),
        "installer_email":       find_right(di, "Contact Email", within=(19, 21)),
        "delivery_address":      find_right(di, "Address", within=(17, 19)),

        # --- order ---
        "order_reason":       find_right(gi, "Order Reason"),
        "platform":           find_right(gi, "Platform"),
        "migration_required": find_right(gi, "Migration required 3g to 4G"),
        "order_date":         find_right(gi, "Order Date"),
        "sales_rep":          find_right(gi, "Sales Representative"),
        "dealer":             find_right(gi, "Dealership Name"),

        # --- commercials ---
        "total_contract_value": _num(find_right(bi, "Total Contract Value")),
        "dealer_commission":    _num(find_right(bi, "Total Dealer Commission")),
        "install_value":        _num(find_right(bi, "Total Install Value")),
        "monthly_payment":      _num(find_right(bi, "Total Monthly Payment")),

        "contract_term_months": _term(bi),
        "lines": _lines(bi),
        "multi_site": _multi_site(wb),
        "vehicles": _vehicles(wb),
    }

    data["ship_to_type"], data["install_required"] = classify_ship_to(data)
    data["acv"], data["acv_note"] = annual_contract_value(data)

    e164, display, note = normalise_phone(data.get("site_contact_phone"))
    data["site_phone_e164"] = e164
    data["site_phone_display"] = display
    data["site_phone_note"] = note

    data["derived_item_name"] = _item_name(data)
    data["_warnings"] = _validate(data)
    return data


def _num(v):
    if v in EMPTY or v is None:
        return None
    try:
        return float(str(v).replace("$", "").replace(",", ""))
    except ValueError:
        return None


# --------------------------------------------------------------------------
# ACV — Annual Contract Value
# --------------------------------------------------------------------------
# ACV is NOT the total contract value. It is total contract value divided by
# the term in YEARS (the eOrder states the term in months). Navtek use it to
# measure new-business targets, so it must only be populated for the order
# reasons that represent new revenue — the board column is even labelled
# "ACV (ATF/NB only)".
#
# Populating it for a renewal or a change of ownership would inflate the
# target figure with revenue that isn't new.

ACV_ORDER_REASONS = {
    "new business",
    "new customer",
    "add-on",          # observed spelling of Add to Fleet on the Qualityvend eOrder
    "add on",
    "add to fleet",
    "atf",
    "upsell",          # Navtek count upsells as new business for target purposes
}


def annual_contract_value(d):
    """Return (acv, note). acv is None when it should not be populated."""
    reason = (d.get("order_reason") or "").strip().lower()
    if reason not in ACV_ORDER_REASONS:
        return None, f"not applicable — order reason is {d.get('order_reason')!r}"

    tcv = d.get("total_contract_value")
    term = d.get("contract_term_months")
    if not tcv:
        return None, "no total contract value in the eOrder"
    if not term:
        return None, "no contract term found — cannot annualise"
    if term < 12:
        return None, f"term is {term} months — too short to annualise, check manually"

    return round(tcv / (term / 12), 2), None


def _term(bi):
    """Contract term in months.

    Taken as the longest term across real product lines. 'Internal Use Only'
    (Commissions & Fees) and 'Parts' lines carry their own unrelated terms —
    Cosmo's parts line is term 1 — so they must not drive the contract term.
    """
    hdr = None
    for row in bi.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() == "Quantity":
                hdr = c.row
                break
        if hdr:
            break
    if not hdr:
        return None

    terms = []
    for row in bi.iter_rows(min_row=hdr + 1, max_row=hdr + 40):
        qty, name, family, term = row[0].value, row[1].value, row[4].value, row[5].value
        if not isinstance(qty, (int, float)) or not name:
            continue
        fam = str(family or "").strip().lower()
        if fam in ("internal use only", "parts"):
            continue
        if isinstance(term, (int, float)) and term:
            terms.append(int(term))
    return max(terms) if terms else None


def _lines(bi):
    """Product lines. Header row is located by label, not assumed to be row 45."""
    out = []
    header = None
    for row in bi.iter_rows():
        for c in row:
            if isinstance(c.value, str) and c.value.strip() == "Quantity":
                header = c.row
                break
        if header:
            break
    if not header:
        return out

    for row in bi.iter_rows(min_row=header + 1, max_row=header + 40):
        qty, name = row[0].value, row[1].value
        if qty in EMPTY or name in EMPTY:
            continue
        if not isinstance(qty, (int, float)):
            continue
        if "Commissions" in str(name):          # internal-use line, not hardware
            continue
        out.append({"qty": int(qty), "product": str(name).strip()})
    return out


def _multi_site(wb):
    """Rows from 'Multi Delivery&Install Details' — one install site each.

    Populated only for multi-city orders (e.g. JETS: 4 MEL, 8 SYD, 2 BNE, 1 ADL).
    Each row should become one subitem on the monday order, with its own
    contact and its own SLA clock.
    """
    ws = wb.get("Multi Delivery&Install Details") if hasattr(wb, "get") else None
    if ws is None:
        if "Multi Delivery&Install Details" not in wb.sheetnames:
            return []
        ws = wb["Multi Delivery&Install Details"]

    out = []
    header = None
    for row in ws.iter_rows():
        vals = [str(c.value).strip() if c.value else "" for c in row]
        if "Company Name" in vals and "Contact Name" in vals:
            header = row[0].row
            break
    if not header:
        return out

    for row in ws.iter_rows(min_row=header + 1, max_row=header + 60):
        v = [c.value for c in row[:6]]
        if all(x in EMPTY for x in v):
            continue
        # Every eOrder sheet ends with a "© 2022 Teletrac Navman" footer that
        # otherwise reads as a site with a company name and nothing else.
        if v[0] and str(v[0]).strip().startswith("©"):
            continue
        if all(x in EMPTY for x in v[1:]):      # a name with no contact detail
            continue
        e164, disp, note = normalise_phone(v[2])
        out.append({
            "company":  v[0], "contact": v[1],
            "phone": disp, "phone_e164": e164, "phone_note": note,
            "email":    v[3], "address": v[4], "notes":   v[5],
            "qty": site_qty(v[5]),
        })
    return out


# The per-site quantity lives in the free-text notes column and the format is
# NOT consistent between eOrders:
#   Qantas       -> 'x 4'
#   Qualityvend  -> '8 x VT202s get shipped to Paul (Sydney installs)'
# So this reads the first quantity-looking token either side of an 'x'. Returns
# None if it can't tell, which must NOT be treated as zero — a failed read and a
# genuine zero mean very different things when reconciling against line totals.
QTY_PATTERNS = [
    re.compile(r"^\s*x\s*(\d+)", re.I),      # 'x 4'
    re.compile(r"^\s*(\d+)\s*x", re.I),      # '8 x VT202s ...'
    re.compile(r"\bx\s*(\d+)\b", re.I),      # '... x 4 ...'
]


def site_qty(notes):
    if not notes:
        return None
    text = str(notes)
    for pat in QTY_PATTERNS:
        m = pat.search(text)
        if m:
            return int(m.group(1))
    return None


# TN ships the Onboarding Vehicles sheet pre-filled with red example text and
# tells the user to delete it. Nobody does. These placeholder rows appeared
# IDENTICALLY in all five sample eOrders, including ones with no vehicles at
# all — so anything matching them must be discarded or you will publish a
# Kenworth K900 that does not exist to an installer's job sheet.
TEMPLATE_VEHICLES = {"1jg3jk", "bin 1", "bin1"}
TEMPLATE_VIN = "65F45678643132"


def _vehicles(wb):
    """Rego / VIN / make / model per vehicle.

    NOTE: across all five validation files this returned ONLY template
    placeholder data — i.e. the team is not currently filling this sheet in,
    despite step 4 of the Salesforce process asking them to. Treat an empty
    result as normal, not as a parse failure.
    """
    if "Onboarding Vehicles" not in wb.sheetnames:
        return []
    ws = wb["Onboarding Vehicles"]
    out = []
    header = None
    for row in ws.iter_rows(max_row=30):
        vals = [str(c.value).strip() if c.value else "" for c in row]
        if "Registration" in vals and "VIN" in vals:
            header = row[0].row
            break
    if not header:
        return out

    for row in ws.iter_rows(min_row=header + 1, max_row=header + 200):
        rego = row[5].value if len(row) > 5 else None   # column F = Rego
        if rego in EMPTY:
            continue
        vin = str(row[6].value).strip() if row[6].value not in EMPTY else None
        if str(rego).strip().lower() in TEMPLATE_VEHICLES or vin == TEMPLATE_VIN:
            continue                                    # template example row
        out.append({
            "rego":  str(rego).strip(),
            "vin":   vin,
            "make":  str(row[8].value).strip() if row[8].value not in EMPTY else None,
            "model": str(row[9].value).strip() if row[9].value not in EMPTY else None,
        })
    return out


# --------------------------------------------------------------------------
# Ship-to classification
# --------------------------------------------------------------------------
# The "Company Name" under Delivery Address is the SHIP-TO, not necessarily the
# installer. Observed values across the samples:
#
#   FFT TECHNOLOGY          -> a real installer; contact name + email present
#   Cosmo Cranes Pty Ltd    -> shipping to the customer themselves
#   Nothing to Ship         -> no hardware, therefore no install
#   Multiple Addresses      -> see the Multi Delivery&Install Details sheet
#
# Only the first case yields an installer. Do not assume the eOrder names one.

SHIP_NOTHING = "nothing to ship"
SHIP_MULTIPLE = "multiple addresses"

# Legal-entity suffixes stripped before comparing two company names.
# NOTE: this must be an anchored regex, not str.rstrip(). rstrip() takes a SET
# of characters, so rstrip(' pty ltd') reduces 'Ally Pty Ltd' to 'a' and
# 'Kane Civil Pty Ltd' to 'kane civi'. Since this function decides the
# Install Required? column, a false match there is a real-world wrong answer.
ENTITY_SUFFIX = re.compile(
    r"\s*(?:"
    r"p\s*/\s*l\.?"                         # P/L, P / L
    r"|pty\.?\s*l(?:td|imited)?\.?"         # Pty Ltd, Pty. Ltd., Pty
    r"|pty\.?"
    r"|l(?:td|imited)\.?"
    r"|inc\.?|llc\.?|nl\.?"
    r"|(?:a|the)?\s*trust(?:ee)?(?:\s+for.*)?"
    r"|t/a\s+.*"                            # trading-as tail
    r")\s*$",
    re.I,
)


def company_key(name):
    """Comparable form of a company name: suffix stripped, punctuation and
    whitespace normalised, lowercased. Applied repeatedly so 'X Pty Ltd.' and
    'X Pty' both reduce to 'x'."""
    s = str(name or "").strip()
    for _ in range(3):
        new = ENTITY_SUFFIX.sub("", s).strip(" .,-")
        if new == s:
            break
        s = new
    return " ".join(s.lower().split())


def classify_ship_to(data):
    """Return (ship_to_type, install_required) — install_required matches the
    monday 'Install Required?' status column."""
    company = (data.get("installer_company") or "").strip()
    low = company.lower()

    if not company:
        return "unknown", None
    if low == SHIP_NOTHING:
        return "nothing_to_ship", "No"
    if low == SHIP_MULTIPLE:
        return "multiple", "Yes"
    if company_key(company) and company_key(company) == company_key(data.get("company")):
        return "customer", "Customer self-install"
    if data.get("installer_contact"):
        return "installer", "Yes"
    return "third_party", "Yes"


# --------------------------------------------------------------------------
# Phone normalisation
# --------------------------------------------------------------------------
# Observed in the samples: '0437353834', '0409 900 543 ' (trailing space),
# '(040) 728-3080' (US formatting of an AU mobile) and '1300 1 COSMO'
# (alphanumeric vanity number). The portal needs tel: links and the SMS step
# needs E.164, so both need handling.

VANITY = str.maketrans("ABCDEFGHIJKLMNOPQRSTUVWXYZ", "22233344455566677778889999")


def normalise_phone(raw, default_country="61"):
    """Best-effort AU normalisation. Returns (e164_or_None, display, note).

    Handles the case where Excel has stored the number as a NUMBER and eaten the
    leading zero — seen in the Qualityvend multi-site sheet, where Paul Redmond's
    mobile arrives as the integer 411477304 rather than the string '0411477304'.
    Build a tel: link from that unrepaired and the call fails.
    """
    if raw in (None, ""):
        return None, None, "no phone supplied"

    if isinstance(raw, (int, float)):                 # Excel ate the leading zero
        raw = str(int(raw))
        if len(raw) == 9 and raw[0] in "2345678":
            raw = "0" + raw

    display = " ".join(str(raw).split())

    letters = re.sub(r"[^A-Za-z]", "", display)
    work = display.upper().translate(VANITY) if letters else display
    digits = re.sub(r"\D", "", work)

    note = None
    if letters:
        note = f"vanity number converted: {display}"

    if digits.startswith("61"):
        digits = "0" + digits[2:]
    if len(digits) == 10 and digits.startswith("0"):
        return f"+{default_country}{digits[1:]}", display, note
    if len(digits) in (6, 10, 11) and digits.startswith("13"):
        return f"+{default_country}{digits}", display, note or "13/1300 number"
    return None, display, (note or "could not normalise — check manually")


PRODUCT_CODE = re.compile(r"^([A-Z]{2,4})\s?(\d{3,4})")


def product_code(name):
    """'RE 400 with TN360' -> 'RE400';  'VT202 VLU with TN360' -> 'VT202'.

    TN writes the same product with and without a space in the code, which is
    why splitting on whitespace alone produces 'RE'.
    """
    name = name.strip()
    m = PRODUCT_CODE.match(name)
    if m:
        return m.group(1) + m.group(2)
    # Descriptive products with no code ('Heavy Vehicle Camera Bundle 4 - Dual
    # Camera and 2 Side Cameras 512GB...'). Cut at the first ' - ' or ' with ',
    # which is where TN starts the long description.
    short = re.split(r"\s+-\s+|\s+with\s+", name)[0]
    return " ".join(short.split()[:5])


def _item_name(d):
    """Rebuild Navtek's existing board naming convention:

        KANE CIVIL = 18 x RE400, 22 x VT202, 4 x AT551

    Company is UPPERCASED with the legal suffix removed, matching the existing
    rows on the board. Do not split on a literal ' PTY' — eOrders write the
    suffix in mixed case ('Cosmo Cranes Pty Ltd'), so a case-sensitive split
    silently leaves it in for most customers while passing on Kane Civil.
    """
    if not d.get("company"):
        return None
    company = company_key(d["company"]).upper()
    parts = [f"{l['qty']} x {product_code(l['product'])}" for l in d["lines"]]
    return f"{company} = " + ", ".join(parts) if parts else company


# --------------------------------------------------------------------------
# Validation — decides ✅ vs ⚠️ on the monday row
# --------------------------------------------------------------------------

CRITICAL = ["opportunity_id", "company"]
EXPECTED = ["site_contact_name", "site_contact_phone"]


def _validate(d):
    warnings = []
    for f in CRITICAL:
        if not d.get(f):
            warnings.append(f"CRITICAL: {f} missing — cannot match to a monday row")
    for f in EXPECTED:
        if not d.get(f):
            warnings.append(f"missing: {f}")

    # Installer allocation. NOTE: the eOrder names an installer only when the
    # hardware ships to one. It is NOT an error for it to be absent — Navtek
    # allocates the installer separately in most cases.
    ship = d.get("ship_to_type")
    if ship == "multiple":
        if not d.get("multi_site"):
            warnings.append("ship-to is 'Multiple Addresses' but the multi-site sheet is empty")
        else:
            sites = d["multi_site"]
            warnings.append(
                f"multi-site order: {len(sites)} delivery sites — "
                "create one subitem per site, each with its own SLA clock")

            # Reconcile per-site quantities against the line-item total, but only
            # when every site's quantity was actually readable from free text.
            qtys = [s.get("qty") for s in sites]
            total = sum(l["qty"] for l in d.get("lines", []))
            if any(q is None for q in qtys):
                warnings.append(
                    "could not read a quantity from every site's notes — "
                    "reconcile the split manually")
            elif total and sum(qtys) != total:
                warnings.append(
                    f"site quantities sum to {sum(qtys)} but the order is for {total}")
    elif ship == "third_party":
        warnings.append("ships to a third party but no contact name given — installer needs allocating")
    elif ship == "unknown":
        warnings.append("ship-to company blank — cannot tell whether an install is needed")

    if d.get("site_phone_note"):
        warnings.append(f"phone: {d['site_phone_note']}")

    if not d.get("lines"):
        warnings.append("no product lines found — check the Billing sheet layout")

    return warnings


# --------------------------------------------------------------------------

if __name__ == "__main__":
    result = parse(sys.argv[1])
    print(json.dumps(result, indent=2, default=str))
    if result["_warnings"]:
        print("\nWARNINGS:", file=sys.stderr)
        for w in result["_warnings"]:
            print("  -", w, file=sys.stderr)
